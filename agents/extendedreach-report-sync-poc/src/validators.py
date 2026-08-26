"""File validation and naming.

Validation runs before upload and is the only gate: nothing reaches Google
Drive that has not passed every check configured for it.

The checks report *categories*, never file contents. A validator that echoed
the offending row into a log would defeat the point of the log redaction.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Sequence

FILENAME_TIMESTAMP_FORMAT = "%Y-%m-%d_%H%M%S"
RUN_KEY_DATE_FORMAT = "%Y-%m-%d"

APPROVED_EXTENSIONS = ("csv", "xlsx", "xls", "pdf")

# Error categories. Fixed vocabulary — these are what reaches the log.
CATEGORY_MISSING_FILE = "file_missing"
CATEGORY_TOO_SMALL = "file_too_small"
CATEGORY_BAD_EXTENSION = "extension_not_allowed"
CATEGORY_CSV_HEADERS = "csv_headers_missing"
CATEGORY_CSV_UNREADABLE = "csv_unreadable"
CATEGORY_XLSX_UNREADABLE = "xlsx_unreadable"
CATEGORY_XLS_UNREADABLE = "xls_unreadable"
CATEGORY_PDF_UNREADABLE = "pdf_unreadable"

_SLUG_CLEAN = re.compile(r"[^a-z0-9]+")


def slugify(value: str) -> str:
    """Lowercase, underscore-separated, safe in a filename.

    Report slugs come from configuration, but a slug is interpolated into a
    path, so it is normalised rather than trusted.
    """
    cleaned = _SLUG_CLEAN.sub("_", (value or "").strip().lower()).strip("_")
    return cleaned or "report"


def normalise_extension(value: str) -> str:
    return (value or "").strip().lstrip(".").lower()


def build_filename(report_slug: str, extension: str,
                   when: Optional[datetime] = None) -> str:
    """extendedreach_<report_slug>_<YYYY-MM-DD_HHMMSS>.<extension>"""
    when = when or datetime.now()
    stamp = when.strftime(FILENAME_TIMESTAMP_FORMAT)
    return f"extendedreach_{slugify(report_slug)}_{stamp}.{normalise_extension(extension)}"


def build_run_key(report_slug: str, when: Optional[datetime] = None) -> str:
    """Deterministic key for one report on one calendar day.

    Two runs of the same report on the same day produce the same key, which is
    what makes the duplicate check work even though the filename carries a
    time. Change the granularity here if a report is exported more than once a
    day on purpose.
    """
    when = when or datetime.now()
    return f"extendedreach:{slugify(report_slug)}:{when.strftime(RUN_KEY_DATE_FORMAT)}"


def filename_prefix_for_run_key(report_slug: str,
                                when: Optional[datetime] = None) -> str:
    """The filename prefix shared by every file with the same run key.

    Used as the fallback duplicate check when a previous upload predates
    appProperties, or was placed in the folder by hand.
    """
    when = when or datetime.now()
    return f"extendedreach_{slugify(report_slug)}_{when.strftime(RUN_KEY_DATE_FORMAT)}"


@dataclass
class ValidationResult:
    ok: bool
    category: Optional[str] = None
    detail: str = ""
    checks: list[str] = field(default_factory=list)

    def __bool__(self) -> bool:
        return self.ok


def _fail(category: str, detail: str, checks: list[str]) -> ValidationResult:
    return ValidationResult(False, category, detail, checks)


def read_csv_header(path: Path,
                    encodings: Sequence[str] = ("utf-8-sig", "cp1252")) -> list[str]:
    """First row of a CSV, trying each encoding in turn.

    Not every ExtendedReach CSV is UTF-8 — the Compliance export is
    Windows-1252 — so a single-encoding read would fail on a valid file.
    """
    last_error: Optional[Exception] = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                sample = handle.readline()
                if not sample:
                    return []
                reader = csv.reader(io.StringIO(sample))
                return [cell.strip().lstrip("﻿") for cell in next(reader, [])]
        except (UnicodeDecodeError, csv.Error, StopIteration) as exc:
            last_error = exc
            continue
    if last_error:
        raise ValueError("header row unreadable in every configured encoding")
    return []


def missing_headers(found: Iterable[str], expected: Iterable[str]) -> list[str]:
    """Expected headers absent from the file, compared case-insensitively and
    ignoring surrounding whitespace. Extra columns are allowed: reports gain
    columns over time and that should not fail a run."""
    normalised = {h.strip().lower() for h in found if h is not None}
    return [e for e in expected if e.strip().lower() not in normalised]


def validate_xlsx(path: Path) -> bool:
    """True if the workbook opens and has at least one sheet."""
    try:
        from openpyxl import load_workbook
    except ImportError:                                  # pragma: no cover
        # Without openpyxl, fall back to the container check: xlsx is a zip.
        return zipfile.is_zipfile(path)
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return bool(workbook.sheetnames)
        finally:
            workbook.close()
    except Exception:
        return False


def validate_xls(path: Path) -> bool:
    """Legacy .xls is an OLE2 compound file. Check the magic bytes rather than
    taking a dependency on xlrd, which no longer reads modern files anyway."""
    try:
        with path.open("rb") as handle:
            return handle.read(8) == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
    except OSError:
        return False


def validate_pdf(path: Path) -> bool:
    try:
        with path.open("rb") as handle:
            return handle.read(5) == b"%PDF-"
    except OSError:
        return False


def validate_file(path: Path,
                  *,
                  min_bytes: int = 1024,
                  allowed_extensions: Iterable[str] = APPROVED_EXTENSIONS,
                  expected_csv_headers: Sequence[str] = (),
                  csv_encodings: Sequence[str] = ("utf-8-sig", "cp1252"),
                  ) -> ValidationResult:
    """Run every applicable check. Stops at the first failure.

    Returns a category, never file content: a missing-header report names the
    *expected* headers (which came from configuration) and never the headers
    actually found, which are file data.
    """
    checks: list[str] = []
    path = Path(path)

    if not path.exists() or not path.is_file():
        return _fail(CATEGORY_MISSING_FILE, "file does not exist", checks)
    checks.append("exists")

    size = path.stat().st_size
    if size < min_bytes:
        return _fail(CATEGORY_TOO_SMALL,
                     f"{size} bytes is below the {min_bytes}-byte minimum", checks)
    checks.append(f"size>={min_bytes}")

    allowed = {normalise_extension(e) for e in allowed_extensions}
    unknown = allowed - set(APPROVED_EXTENSIONS)
    if unknown:
        return _fail(CATEGORY_BAD_EXTENSION,
                     f"configured extensions outside the approved list: {sorted(unknown)}",
                     checks)

    extension = normalise_extension(path.suffix)
    if extension not in allowed:
        return _fail(CATEGORY_BAD_EXTENSION,
                     f"'{extension}' is not in the allow-list {sorted(allowed)}", checks)
    checks.append(f"extension={extension}")

    if extension == "csv":
        try:
            header = read_csv_header(path, csv_encodings)
        except ValueError as exc:
            return _fail(CATEGORY_CSV_UNREADABLE, str(exc), checks)
        if not header:
            return _fail(CATEGORY_CSV_UNREADABLE, "no header row", checks)
        checks.append(f"csv_header_columns={len(header)}")
        if expected_csv_headers:
            missing = missing_headers(header, expected_csv_headers)
            if missing:
                return _fail(CATEGORY_CSV_HEADERS,
                             f"missing expected column(s): {missing}", checks)
            checks.append(f"csv_expected_headers={len(list(expected_csv_headers))}")

    elif extension == "xlsx":
        if not validate_xlsx(path):
            return _fail(CATEGORY_XLSX_UNREADABLE, "workbook did not open", checks)
        checks.append("xlsx_opens")

    elif extension == "xls":
        if not validate_xls(path):
            return _fail(CATEGORY_XLS_UNREADABLE,
                         "not an OLE2 compound file; the download may be an "
                         "error page saved with an .xls name", checks)
        checks.append("xls_signature")

    elif extension == "pdf":
        if not validate_pdf(path):
            return _fail(CATEGORY_PDF_UNREADABLE, "missing %PDF- signature", checks)
        checks.append("pdf_signature")

    return ValidationResult(True, None, "all checks passed", checks)
